import pandas as pd
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from openpyxl import load_workbook

# 현재 실행 경로
base_dir = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(base_dir, "123.xlsx")
image_path = os.path.join(base_dir, "g.jpg")
pdf_filename = os.path.join(base_dir, "envelopes.pdf")

# 한글 폰트 등록
font_path = "C:/Windows/Fonts/H2GTRE.TTF"
pdfmetrics.registerFont(TTFont("H2GTRE", font_path))

# 엑셀 데이터 불러오기
df = pd.read_excel(file_path)
df.columns = df.columns.str.strip()

# openpyxl 로드 (색상용)
wb = load_workbook(file_path, data_only=True)
ws = wb.active

# 색상 추출 함수
def get_rgb_color(cell):
    font_color = cell.font.color
    if font_color and font_color.type == 'rgb' and font_color.rgb:
        rgb = font_color.rgb[-6:]
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
        return (r / 255.0, g / 255.0, b / 255.0)
    return (0, 0, 0)  # 기본 검정

# 봉투 크기 설정
mm_to_pt = 2.8346457
envelope_width = 220 * mm_to_pt
envelope_height = 110 * mm_to_pt

# PDF 생성
c = canvas.Canvas(pdf_filename, pagesize=(envelope_width, envelope_height))

# 공통 설정
font_size = 18
start_x = 100
start_y = envelope_height - 230
line_spacing = 30

# 로고 설정
logo_size = (100, 100)
logo_position = (envelope_width - 100, envelope_height - 100)
brand_position = (envelope_width - 90, envelope_height - 85)

for idx, row in df.iterrows():
    c.setFont("H2GTRE", font_size)

    # 로고 삽입
    if os.path.exists(image_path):
        logo = ImageReader(image_path)
        c.drawImage(logo, logo_position[0], logo_position[1], width=logo_size[0], height=logo_size[1], mask='auto')

    # 브랜드명
    c.setFont("H2GTRE", 18)
    c.setFillColorRGB(0, 0, 0)
    c.drawRightString(brand_position[0] - 20, brand_position[1] + 45, "기린")
    c.drawRightString(brand_position[0], brand_position[1] + 10, "(길라인)")

    # 셀 색상 추출
    store_cell = ws.cell(row=idx + 2, column=df.columns.get_loc("상가명") + 1)
    biz_cell = ws.cell(row=idx + 2, column=df.columns.get_loc("상호") + 1)
    price_cell = ws.cell(row=idx + 2, column=df.columns.get_loc("금액") + 1)

    store_color = get_rgb_color(store_cell)
    biz_color = get_rgb_color(biz_cell)
    price_color = get_rgb_color(price_cell)

    store_name = str(row["상가명"])
    business_name = str(row["상호"])
    amount = row["금액"]

    # 금액 쉼표 포맷 적용
    if isinstance(amount, (int, float)):
        amount_str = f"{amount:,.0f}원"
    else:
        amount_str = str(amount)

    # ✅ 한 줄에 상가명 → 상호 → 금액 순으로, 위치 자동 조절
    x = start_x

    # 상가명
    c.setFillColorRGB(*store_color)
    c.drawString(x, start_y, store_name)
    store_width = stringWidth(store_name, "H2GTRE", font_size)
    x += store_width + 30

    # 상호
    c.setFillColorRGB(*biz_color)
    c.drawString(x, start_y, business_name)
    biz_width = stringWidth(business_name, "H2GTRE", font_size)
    x += biz_width + 30

    # 금액
    c.setFillColorRGB(*price_color)
    c.drawString(x, start_y, amount_str)

    c.showPage()

c.save()
print(f"✅ PDF 파일 생성 완료: {pdf_filename}")
