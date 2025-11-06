import streamlit as st
import pandas as pd
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase.pdfmetrics import stringWidth
from openpyxl import load_workbook
from openpyxl.styles import Font as XLFont
import io
import tempfile

# 페이지 설정
st.set_page_config(
    page_title="우편봉투 인쇄 시스템",
    page_icon="📮",
    layout="wide"
)

# 타이틀
st.title("📮 우편봉투 인쇄 시스템")
st.markdown("---")

# 현재 실행 경로
base_dir = os.path.dirname(os.path.abspath(__file__))
number_file_path = os.path.join(base_dir, "number.xlsm")
image_path = os.path.join(base_dir, "g.jpg")

# 한글 폰트 등록
@st.cache_resource
def register_font():
    # 여러 폰트 경로 시도 (한글 지원 폰트)
    font_paths = [
        # Linux (Streamlit Cloud)
        "/usr/share/fonts/truetype/nanum/NanumGothic.ttf",
        "/usr/share/fonts/truetype/nanum/NanumBarunGothic.ttf",
        "/usr/share/fonts/truetype/nanum-coding/NanumGothicCoding.ttf",
        # Windows
        "C:/Windows/Fonts/H2GTRE.TTF",
        "C:/Windows/Fonts/malgun.ttf",
        "C:/Windows/Fonts/gulim.ttc",
        "C:/Windows/Fonts/batang.ttc",
        # macOS
        "/System/Library/Fonts/AppleGothic.ttf",
        "/Library/Fonts/AppleGothic.ttf",
    ]
    
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont("KoreanFont", font_path))
                return True, os.path.basename(font_path)
            except Exception as e:
                continue
    
    # 기본 폰트 사용
    return False, None

font_available, font_name = register_font()

if not font_available:
    FONT_NAME = "Helvetica"
    st.warning("⚠️ 한글 폰트를 찾을 수 없습니다. PDF에 한글이 깨져 보일 수 있습니다.")
else:
    FONT_NAME = "KoreanFont"

# 색상 추출 함수
def get_rgb_color(cell):
    font_color = cell.font.color
    if font_color and font_color.type == 'rgb' and font_color.rgb:
        rgb = font_color.rgb[-6:]
        r = int(rgb[0:2], 16)
        g = int(rgb[2:4], 16)
        b = int(rgb[4:6], 16)
        return (r / 255.0, g / 255.0, b / 255.0)
    return (0, 0, 0)  # 기본 검정

# 데이터 정렬 함수
def sort_data_by_number_file(uploaded_df):
    """업로드된 데이터를 number.xlsm 기준으로 정렬"""
    if not os.path.exists(number_file_path):
        st.error(f"❌ {number_file_path} 파일을 찾을 수 없습니다.")
        return None
    
    # number.xlsm 불러오기
    df_number = pd.read_excel(number_file_path)
    df_number.columns = df_number.columns.str.strip()
    
    # 컬럼명 확인 및 정리
    if uploaded_df.columns[0].startswith('Unnamed'):
        # 첫 행이 실제 헤더인 경우
        uploaded_df.columns = uploaded_df.iloc[0]
        uploaded_df = uploaded_df[1:].reset_index(drop=True)
    
    uploaded_df.columns = uploaded_df.columns.str.strip()
    
    # 상호 컬럼 찾기
    business_col = None
    for col in uploaded_df.columns:
        if '상호' in str(col):
            business_col = col
            break
    
    if business_col is None:
        st.error("❌ 업로드된 파일에서 '상호' 컬럼을 찾을 수 없습니다.")
        return None
    
    # 금액 컬럼 찾기
    amount_col = None
    for col in uploaded_df.columns:
        if '금액' in str(col) or '입금' in str(col):
            amount_col = col
            break
    
    if amount_col is None:
        st.error("❌ 업로드된 파일에서 '금액' 컬럼을 찾을 수 없습니다.")
        return None
    
    # 원본 파일에 상가명 컬럼이 있는지 확인
    original_brand_col = None
    for col in uploaded_df.columns:
        if '상가' in str(col):
            original_brand_col = col
            break
    
    # number.xlsm의 컬럼 확인
    brand_col = df_number.columns[0]  # 브랜드/상가명
    number_business_col = df_number.columns[1]  # 상호
    order_col = df_number.columns[2]  # 순서
    
    # 데이터 병합
    merged_df = uploaded_df.merge(
        df_number[[brand_col, number_business_col, order_col]],
        left_on=business_col,
        right_on=number_business_col,
        how='left'
    )
    
    # 원본 파일에 상가명이 있으면 매칭 안 된 경우 원본 상가명 사용
    if original_brand_col:
        merged_df[brand_col] = merged_df[brand_col].fillna(merged_df[original_brand_col])
    
    # 매칭 여부 확인 (순서번호가 있으면 매칭된 것)
    merged_df['has_order'] = merged_df[order_col].notna()
    
    # number.xlsm에 있는 모든 상가명 목록
    all_brands_in_number = df_number[brand_col].unique()
    
    # 정렬을 위한 키 생성
    def get_sort_key(row):
        brand = row[brand_col] if pd.notna(row[brand_col]) else ""
        has_order = row['has_order']
        order_num = row[order_col] if pd.notna(row[order_col]) else 999999
        
        # 해당 상가가 number.xlsm에 존재하는지 확인
        brand_exists_in_number = brand in all_brands_in_number
        
        if not brand_exists_in_number:
            # number.xlsm에 아예 없는 상가 → 맨 앞 (0)
            return (0, brand, 0, 0)
        elif has_order:
            # number.xlsm에 있고 순서번호도 있음 → 중간 (1)
            return (1, brand, 0, order_num)
        else:
            # number.xlsm에 상가는 있지만 이 상호는 없음 → 해당 상가의 뒤 (1, brand, 1)
            return (1, brand, 1, 999999)
    
    merged_df['sort_key'] = merged_df.apply(get_sort_key, axis=1)
    merged_df = merged_df.sort_values('sort_key').reset_index(drop=True)
    
    # 상가명 앞에 순서번호 추가
    result_rows = []
    current_brand = None
    brand_counter = 0
    
    for idx, row in merged_df.iterrows():
        brand_name = str(row[brand_col]) if pd.notna(row[brand_col]) else ""
        business_name = str(row[business_col]) if pd.notna(row[business_col]) else ""
        amount = row[amount_col]
        has_order = row['has_order']
        
        # 순서번호가 있는 경우에만 상가명 앞에 번호 추가
        if has_order and brand_name:
            # 새로운 상가가 시작되면 카운터 리셋
            if brand_name != current_brand:
                current_brand = brand_name
                brand_counter = 1
            else:
                brand_counter += 1
            
            # 이미 숫자로 시작하는 경우 그대로 사용
            if brand_name and brand_name[0].isdigit():
                formatted_brand = brand_name
            else:
                formatted_brand = f"{brand_counter}{brand_name}"
        else:
            # 순서번호가 없으면 상가명만 (번호 없이)
            formatted_brand = brand_name
        
        result_rows.append({
            '상가명': formatted_brand,
            '상호': business_name,
            '금액': amount
        })
    
    result_df = pd.DataFrame(result_rows)
    
    return result_df

# PDF 생성 함수
def create_envelopes_pdf(df, extra_text="", text_size=12, text_color=(0, 0, 0)):
    """봉투 PDF 생성"""
    global FONT_NAME
    # 임시 파일 생성
    temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
    pdf_filename = temp_pdf.name
    temp_pdf.close()
    
    # 봉투 크기 설정
    mm_to_pt = 2.8346457
    envelope_width = 220 * mm_to_pt
    envelope_height = 110 * mm_to_pt
    
    # PDF 생성
    c = canvas.Canvas(pdf_filename, pagesize=(envelope_width, envelope_height))
    
    # 공통 설정
    font_size = 18
    start_x = 100
    start_y = envelope_height - 230
    line_spacing = 30
    
    # 로고 설정
    logo_size = (100, 100)
    logo_position = (envelope_width - 100, envelope_height - 100)
    brand_position = (envelope_width - 90, envelope_height - 85)
    
    # 추가 텍스트 위치
    extra_text_y = start_y - 50
    
    for idx, row in df.iterrows():
        c.setFont(FONT_NAME, font_size)
        
        # 로고 삽입
        if os.path.exists(image_path):
            logo = ImageReader(image_path)
            c.drawImage(logo, logo_position[0], logo_position[1], 
                       width=logo_size[0], height=logo_size[1], mask='auto')
        
        # 브랜드명
        c.setFont(FONT_NAME, 18)
        c.setFillColorRGB(0, 0, 0)
        c.drawRightString(brand_position[0] - 20, brand_position[1] + 45, "기린")
        c.drawRightString(brand_position[0], brand_position[1] + 10, "(길라인)")
        
        store_name = str(row["상가명"]) if pd.notna(row["상가명"]) else ""
        business_name = str(row["상호"]) if pd.notna(row["상호"]) else ""
        amount = row["금액"]
        
        # 금액 쉼표 포맷 적용
        if isinstance(amount, (int, float)):
            amount_str = f"{amount:,.0f}원"
        else:
            amount_str = str(amount)
        
        # 한 줄에 상가명 → 상호 → 금액 순으로, 위치 자동 조절
        x = start_x
        
        # 상가명
        c.setFillColorRGB(0, 0, 0)  # 기본 검정색
        c.drawString(x, start_y, store_name)
        store_width = stringWidth(store_name, FONT_NAME, font_size)
        x += store_width + 30
        
        # 상호
        c.setFillColorRGB(0, 0, 0)
        c.drawString(x, start_y, business_name)
        biz_width = stringWidth(business_name, FONT_NAME, font_size)
        x += biz_width + 30
        
        # 금액
        c.setFillColorRGB(0, 0, 0)
        c.drawString(x, start_y, amount_str)
        
        # 추가 텍스트
        if extra_text:
            c.setFont(FONT_NAME, text_size)
            c.setFillColorRGB(text_color[0], text_color[1], text_color[2])
            c.drawString(start_x, extra_text_y, extra_text)
        
        c.showPage()
    
    c.save()
    
    return pdf_filename

# Excel 생성 함수 (색상 포함)
def create_colored_excel(df, original_file=None):
    """색상이 포함된 엑셀 파일 생성"""
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        
        # 색상 적용 (원본 파일에서 추출)
        if original_file is not None:
            try:
                wb_original = load_workbook(original_file, data_only=False)
                ws_original = wb_original.active
                
                wb_new = writer.book
                ws_new = wb_new.active
                
                # 헤더 스타일 적용
                for col_idx, col_name in enumerate(df.columns, start=1):
                    cell = ws_new.cell(row=1, column=col_idx)
                    cell.font = XLFont(bold=True)
                
                # 데이터 행 스타일 적용 (간단한 버전)
                for row_idx in range(2, len(df) + 2):
                    for col_idx in range(1, len(df.columns) + 1):
                        cell = ws_new.cell(row=row_idx, column=col_idx)
                        # 기본 스타일만 적용
            except Exception as e:
                st.warning(f"색상 적용 중 오류: {str(e)}")
    
    output.seek(0)
    return output

# Session State 초기화
if 'sorted_data' not in st.session_state:
    st.session_state.sorted_data = None
if 'excel_data' not in st.session_state:
    st.session_state.excel_data = None
if 'pdf_data' not in st.session_state:
    st.session_state.pdf_data = None

# 메인 UI
col1, col2 = st.columns([2, 1])

with col1:
    st.header("1️⃣ 엑셀 파일 업로드")
    uploaded_file = st.file_uploader(
        "5.xlsx 형식의 엑셀 파일을 업로드하세요",
        type=['xlsx', 'xls'],
        help="상호와 금액 정보가 포함된 엑셀 파일"
    )

with col2:
    st.header("2️⃣ 추가 텍스트 설정")
    extra_text = st.text_input(
        "봉투에 추가할 내용",
        placeholder="예: 감사합니다",
        help="우편봉투에 표시될 추가 텍스트"
    )
    
    text_size = st.slider(
        "글씨 크기",
        min_value=8,
        max_value=30,
        value=12,
        step=1
    )
    
    text_color_hex = st.color_picker(
        "글씨 색상",
        value="#000000"
    )
    
    # HEX를 RGB로 변환
    text_color_rgb = tuple(int(text_color_hex.lstrip('#')[i:i+2], 16) / 255.0 for i in (0, 2, 4))

st.markdown("---")

# 파일이 업로드되면 처리
if uploaded_file is not None:
    try:
        # 업로드된 파일 읽기
        df_uploaded = pd.read_excel(uploaded_file)
        
        st.success("✅ 파일이 성공적으로 업로드되었습니다!")
        
        with st.expander("📊 업로드된 데이터 미리보기"):
            st.dataframe(df_uploaded.head(10))
        
        # 정렬 버튼
        if st.button("🔄 데이터 정렬 및 PDF 생성", type="primary", use_container_width=True):
            with st.spinner("처리 중..."):
                # 데이터 정렬
                sorted_df = sort_data_by_number_file(df_uploaded)
                
                if sorted_df is not None:
                    st.success("✅ 데이터가 성공적으로 정렬되었습니다!")
                    
                    # 엑셀 파일 생성
                    excel_output = create_colored_excel(sorted_df, uploaded_file)
                    
                    # PDF 생성
                    pdf_file = create_envelopes_pdf(
                        sorted_df, 
                        extra_text=extra_text,
                        text_size=text_size,
                        text_color=text_color_rgb
                    )
                    
                    # Session State에 저장
                    st.session_state.sorted_data = sorted_df
                    st.session_state.excel_data = excel_output.getvalue()
                    
                    with open(pdf_file, 'rb') as f:
                        st.session_state.pdf_data = f.read()
                    
                    # 임시 파일 삭제
                    try:
                        os.unlink(pdf_file)
                    except:
                        pass
                    
                    st.success("✅ PDF가 성공적으로 생성되었습니다!")
                    st.rerun()
        
        # 정렬된 데이터가 있으면 표시
        if st.session_state.sorted_data is not None:
            # 정렬된 데이터 미리보기
            with st.expander("📊 정렬된 데이터 미리보기", expanded=True):
                st.dataframe(st.session_state.sorted_data.head(20))
                st.info(f"총 {len(st.session_state.sorted_data)}개의 행이 정렬되었습니다.")
            
            # 다운로드 버튼 (항상 표시)
            col_dl1, col_dl2 = st.columns(2)
            
            with col_dl1:
                st.download_button(
                    label="📥 정렬된 엑셀 다운로드",
                    data=st.session_state.excel_data,
                    file_name="sorted_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="download_excel"
                )
            
            with col_dl2:
                st.download_button(
                    label="📥 우편봉투 PDF 다운로드",
                    data=st.session_state.pdf_data,
                    file_name="envelopes.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                    key="download_pdf"
                )
        
    except Exception as e:
        st.error(f"❌ 오류가 발생했습니다: {str(e)}")
        st.exception(e)

else:
    # 파일 업로드가 없으면 세션 초기화
    st.session_state.sorted_data = None
    st.session_state.excel_data = None
    st.session_state.pdf_data = None
    st.info("👆 엑셀 파일을 업로드하여 시작하세요.")
    
    # 사용 방법 안내
    with st.expander("📖 사용 방법"):
        st.markdown("""
        ### 사용 방법
        
        1. **엑셀 파일 업로드**: 상호와 금액 정보가 포함된 엑셀 파일을 업로드합니다.
        2. **추가 텍스트 설정**: 
           - 봉투에 표시할 추가 텍스트를 입력합니다.
           - 글씨 크기를 조절합니다 (8~30).
           - 글씨 색상을 선택합니다.
        3. **처리**: "데이터 정렬 및 PDF 생성" 버튼을 클릭합니다.
        4. **다운로드**: 정렬된 엑셀 파일과 우편봉투 PDF 파일을 다운로드합니다.
        
        ### 파일 형식
        
        - 업로드 파일: `상호`, `금액` (또는 `입금금액`) 컬럼 필요
        - number.xlsm: 상가명, 상호, 순서 정보 포함
        - 출력: 상가명, 상호, 금액 순으로 정렬된 데이터
        """)

# 푸터
st.markdown("---")
st.markdown(
    "<div style='text-align: center; color: gray;'>우편봉투 인쇄 시스템 v1.0</div>",
    unsafe_allow_html=True
)

